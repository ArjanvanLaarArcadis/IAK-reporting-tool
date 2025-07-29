# -*- coding: utf-8 -*-
"""
This script processes Excel PI Reports that follow from DISK using OpenPyXL and other utilities.
It uses a 'voortgangslijst' to update the configuration variables and populate the PI reports with relevant data.
The voortgangslijst contains information in each row on who did the inspection, who did quality assurance et cetera.

The script performs the following tasks:
- Loads and processes Excel workbooks for PI reports.
- Populates various sheets in the workbook with data and formatting.
- Updates configuration variables based on the workbook and external data.
- Saves the processed workbook and optionally exports it to PDF.  TODO: Does not work as intended.

Key Features:
- Modular functions for populating specific sheets.
- Logging for tracking the processing steps and errors.
- Integration with external utilities for configuration and PDF export.

Dependencies:
- openpyxl: For handling Excel workbooks.
- pandas: For data manipulation.
- src.utils: Custom utility functions for configuration and logging.
- src.get_voortgang: For retrieving progress parameters.
- src.export_excel_to_pdf: For exporting Excel to PDF.

Usage:
Run the script as a standalone program to process all PI reports in the specified batch.

TODO:
- The excel export to pdf is not working as expected
- The sc
"""

# Built-in modules
import os
import math
import logging
import datetime as dt

# External imports
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font
from PIL import JpegImagePlugin

# Local imports
import utils
import utilsxls
from get_voortgang import get_voortgang, get_voortgang_params
from export_excel_to_pdf import run_macro_on_workbook

# Workaround for PIL bug with JpegImagePlugin
JpegImagePlugin._getmp = lambda: None

# Constants for font styles
FONT_ARIAL_7 = Font(name='Arial', size=7, bold=False)
FONT_ARIAL_8 = Font(name='Arial', size=8, bold=False)
FONT_ARIAL_10 = Font(name='Arial', size=10, bold=False)
FONT_ARIAL_10_BOLD = Font(name='Arial', size=10, bold=True)
FONT_ARIAL_12 = Font(name='Arial', size=12, bold=False)
FONT_ARIAL_16 = Font(name='Arial', size=16, bold=False)
FONT_ARIAL_18 = Font(name='Arial', size=18, bold=False)

# Constants for the alignment style
ALIGNMENT_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


def find_inspectierapport(directory: str) -> str:
    """
    Find the most recent file starting with 'inspectieRapport' (case insensitive)
    and ending with '.xlsx' directly in the given directory.

    Parameters:
        directory (str): Directory to search in. Subdirs are not searched.

    Returns:
        str: fullfilename of the most recent matching file if found.
        None: If no matching file is found.
    """
    logging.debug(f"Searching for .xlsx-files in [{directory}], starting with 'inspectieRapport' (case insensitive)")

    # List to hold tuples of (file_path, modification_time)
    matching_files = []

    # Walk through the directory and its subdirectories
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.lower().startswith("inspectierapport") and \
                file.lower().endswith(".xlsx"):
                # Get the full path and modification time of the inspectie
                file_path = os.path.join(root, file)
                file_mtime = os.path.getmtime(file_path)
                matching_files.append((file_path, file_mtime))
                logging.debug(f"Found matching file: [{file_path}]")

    if not matching_files:
        logging.info("No matching file found.")
        return None

    # Select the file name, based on the most recent time
    most_recent_file = max(matching_files, key=lambda x: x[1])[0]
    logging.info(f"Most recent file found: [{most_recent_file}]")
    return most_recent_file  # Full path of the most recent file


def set_footer(
    wb: openpyxl.Workbook, sheet_names: list, variables: dict, sheets_count: int
) -> None:
    """
    Set the footer for all sheets in the workbook.

    Parameters:
        wb (openpyxl.Workbook): The workbook object.
        sheet_names (list): List of sheet names.
        variables (dict): Dictionary of variables.
        sheets_count (int): Total number of sheets.
    """
    logging.debug("Setting footers for all sheets...")
    complex = variables.get("complex_code", "UNKNOWN")
    objectcode = variables.get("object_code", "UNKNOWN")
    versie = variables.get("versie", "UNKNOWN")
    datum = variables.get("datum", "UNKNOWN")
    # object_beheer = variables.get("object_beheer", "UNKNOWN")  # te lang, zorgd voor problemen in de output

    FOOTER_LEFT = f"Complex: {complex}\nBeheerobject: {objectcode}\nVertrouwelijkheid: RWS Bedrijfsvertrouwelijk"
    FOOTER_RIGHT = f"Revisie: {versie}\nDatum: {datum}\nPagina &P van &N"
    for i in range(2, sheets_count):
        sheet = wb[sheet_names[i]]
        list_of_footers = [
            sheet.evenFooter.left,
            sheet.evenFooter.center,
            sheet.evenFooter.right,
            sheet.oddFooter.left,
            sheet.oddFooter.center,
            sheet.oddFooter.right,
        ]
        sheet.evenFooter.left.text = FOOTER_LEFT
        sheet.evenFooter.right.text = FOOTER_RIGHT
        sheet.evenFooter.center.text = ""
        sheet.oddFooter.left.text = FOOTER_LEFT
        sheet.oddFooter.center.text = ""
        sheet.oddFooter.right.text = FOOTER_RIGHT
        for footer in list_of_footers:
            footer.font = "Arial"
            footer.size = 7
    logging.debug("Footers set successfully.")


def populate_title_page(
    sheet: openpyxl.worksheet.worksheet.Worksheet, variables: dict
) -> None:
    """
    Populate and format the Title Page (Sheet2).

    Parameters:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet object.
        variables (dict): Dictionary of variables.
    """
    logging.debug("Populating Title Page (Sheet2)...")

    opdrachtgever = variables.get('opdrachtgever', 'UNKNOWN')
    contactpersoon_rws = variables.get('contactpersoon_rws', 'UNKNOWN')
    zaaknr = variables.get("zaaknummer", "UNKNOWN")
    versie = variables.get('versie', 'UNKNOWN')
    datum = variables.get('datum', 'UNKNOWN')
    omschrijving = variables.get('omschrijving', 'UNKNOWN')
    opdrachtnemer = variables.get('opdrachtnemer', 'UNKNOWN')
    opsteller = variables.get('opsteller', 'UNKNOWN')
    kwaliteitsbeheerder = variables.get("kwaliteitsbeheerser", "UNKNOWN")
    projectleider = variables.get('projectleider', 'UNKNOWN')

    sheet['H14'] = opdrachtgever
    sheet['H15'] = contactpersoon_rws
    sheet['H16'] = str(zaaknr)
    sheet['F23'] = versie
    sheet['J23'] = datum
    sheet['L23'] = omschrijving
    sheet['D25'] = opdrachtnemer
    sheet['C27'] = opsteller
    sheet['I27'] = kwaliteitsbeheerder
    sheet['O27'] = projectleider

    # Adjust row heights
    sheet.row_dimensions[10].height = 30
    sheet.row_dimensions[14].height = 30
    sheet.row_dimensions[27].height = 45

    # Adjust column widths to get the names to fit in the signatures box
    sheet.column_dimensions["C"].width += 1  # Opsteller
    sheet.column_dimensions["L"].width += 2  # Kwaliteitsbeheerder
    sheet.column_dimensions["S"].width += 3  # Projectleider

    sheet.column_dimensions["H"].width -= 2  # Opsteller
    sheet.column_dimensions["N"].width -= 2  # Kwaliteitsbeheerder
    sheet.column_dimensions["V"].width -= 2  # Projectleider

    logging.debug("Title Page populated and formatted successfully.")


def populate_inhoud_sheet(
    sheet: openpyxl.worksheet.worksheet.Worksheet, variables: dict
) -> None:
    """
    Populate and format the Inhoud (Sheet3).

    Parameters:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet object.
        variables (dict): Dictionary of variables.
    """
    logging.debug("Populating Inhoud (Sheet3)...")
    sheet["C5"].font = FONT_ARIAL_10
    sheet.row_dimensions[6].height = 235
    sheet.row_dimensions[7].height = 15
    logging.debug("Inhoud populated and formatted successfully.")


def populate_aanbeveling_sheet(
    sheet: openpyxl.worksheet.worksheet.Worksheet, variables: dict
) -> None:
    """
    Populate and format the Aanbeveling (Sheet4).

    Parameters:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet object.
        variables (dict): Dictionary of variables.
    """
    logging.debug("Populating Aanbeveling (Sheet4)...")
    venr = variables.get("venr")
    nader_onderzoek = variables.get("nader_onderzoek")
    directe_maatregel = variables.get("directe_maatregel")
    sheet["C5"] = "1  Aanbeveling"
    sheet["C5"].font = FONT_ARIAL_18
    # standaard tekst
    intro_text = "Voor de volledige uitwerking van het planoverzicht in MIOK, wordt verwezen naar hoofdstuk 2. Hierin zijn per instandhoudingsonderdeel de onderhouds- en/of vervangingsmaatregelen inclusief bijbehorende kosten opgenomen."
    venr_text = f"V&R-indicatie:\nDe V&R-indicatie geeft weer in welk jaar de verwachte renovatie of vervanging gepland staat. Op basis van de ORA wordt hiervan een inschatting gemaakt. Voor dit object is de V&R-indicatie gesteld op {venr}. Hierbij dient uit te worden gegaan van een volledige renovatie van het object."
    geen_nader_onderzoek_text = "Nader onderzoek:\nZowel vanuit de opgestelde (i-)ORA als vanuit de uitgevoerde inspectiewerkzaamheden is geen noodzaak gebleken voor het uitvoeren van een nader onderzoek."
    wel_nader_onderzoek_text = f"Nader onderzoek:\nVanuit de uitgevoerde inspectiewerkzaamheden wordt het volgende nader onderzoek geadviseerd: {nader_onderzoek}"
    geen_directe_maatregel_text = "Directe maatregelen:\nBij dit object zijn geen directe maatregelen noodzakelijk geacht."
    wel_directe_maatregel_text = f"Directe maatregelen:\nBij dit object zijn de volgende directe maatregelen noodzakelijk geacht: {directe_maatregel}"
    aandachtspunten_beheerder_text = "Aandachtspunten voor de beheerder:\nIn Bijlage 9: Aandachtspunten voor de beheerder zijn de schades opgenomen die geconstateerd zijn tijdens de inspectie maar, volgens de risicoanalyse, geen risico initiëren voor het functioneren van het object. Daarnaast zijn de schades opgenomen die vallen onder standaard verzorgend onderhoud"

    max_row_count = sheet.max_row
    start_table = 9
    for row in sheet[f"D{start_table}:D{max_row_count}"]:
        cell = row[0]
        if cell.value:
            index_row = cell.row
            sheet.row_dimensions[index_row].height = 30
        else:
            end_table = index_row
            break

    counter = 0
    text_voor_aanbeveling = ("\n\n").join(
        [
            intro_text,
            # venr_text,
            (
                wel_nader_onderzoek_text
                if nader_onderzoek
                else geen_nader_onderzoek_text
            ),
            (
                wel_directe_maatregel_text
                if directe_maatregel
                else geen_directe_maatregel_text
            ),
            aandachtspunten_beheerder_text,
        ]
    )

    for row in sheet[f"D{end_table + 1}:D{max_row_count}"]:
        cell = row[0]
        if cell.value and counter == 0:
            counter += 1
            cell.value = text_voor_aanbeveling
            sheet.row_dimensions[cell.row].height = 300
        else:
            cell.value = ""
    logging.debug("Aanbeveling populated and formatted successfully.")


def populate_ihp_sheet(
    sheet: openpyxl.worksheet.worksheet.Worksheet, variables: dict
) -> None:
    """
    Populates the IHP sheet with the given variables.

    Args:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet to populate.
        variables (dict): A dictionary containing the variables to populate the sheet with.

    Returns:
        None
    """
    logging.debug("Populating IHP (Sheet5)...")
    sheet['D4'].font = FONT_ARIAL_18
    logging.debug("IHP populated and formatted successfully.")


def populate_miok_sheet(
    sheet: openpyxl.worksheet.worksheet.Worksheet, variables: dict
) -> None:
    """
    Populate and format the MIOK sheet.

    Parameters:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet object.
        variables (dict): Dictionary of variables.
    """
    logging.debug("Populating MIOK (Sheet6)...")
    row_count = sheet.max_row
    sheet.column_dimensions['AH'].width = 4

    for row in sheet.iter_rows(min_row=1, max_row=row_count):
        for cell in row:
            cell.font = FONT_ARIAL_7

    sheet.print_area = f"A1:AL{row_count}"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToHeight = False
    logging.debug("MIOK populated and formatted successfully.")


def populate_inleiding_sheet(
    sheet: openpyxl.worksheet.worksheet.Worksheet, variables: dict
) -> None:
    """
    Populate and format the Inleiding (Sheet7).

    Parameters:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet object.
        variables (dict): Dictionary of variables.
    """
    logging.debug("Populating Inleiding (Sheet7)...")
    opdrachtnemer = variables.get('opdrachtnemer', 'UNKNOWN')
    inspecteurs = variables.get('inspecteurs', 'UNKNOWN')
    # hulpmiddelen = variables.get('hulpmiddelen', 'UNKNOWN')  # Not used

    sheet['C4'] = '3  Inleiding'
    sheet['C4'].font = FONT_ARIAL_18

    # standaard tekst 3.1
    sheet['C6'].font = FONT_ARIAL_12
    sheet['D6'].font = FONT_ARIAL_12
    sheet['D7'] = f'In opdracht van Rijkswaterstaat is door {opdrachtnemer} een instandhoudingsinspectie uitgevoerd aan'
    sheet['E12'].value = inspecteurs
    sheet.row_dimensions[12].height = 70
    sheet.row_dimensions[13].height = 15
    # sheet['E13'] = hulpmiddelen

    # standaard tekst 3.2
    sheet["D15"].font = FONT_ARIAL_10
    sheet["D15"].alignment = ALIGNMENT_LEFT
    sheet.row_dimensions[15].height = 70
    sheet.row_dimensions[16].height = 15

    # standaard tekst 3.3
    sheet["D18"].font = FONT_ARIAL_10
    sheet["D18"].alignment = ALIGNMENT_LEFT
    sheet.row_dimensions[18].height = 70
    sheet.row_dimensions[19].height = 15

    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToHeight = False
    logging.debug("Inleiding populated and formatted successfully.")


def populate_objectgegevens_sheet(
    sheet: openpyxl.worksheet.worksheet.Worksheet, variables: dict
) -> None:
    """
    Populate and format the Objectgegevens (Sheet8).

    Parameters:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet object.
        variables (dict): Dictionary of variables.
    """
    logging.debug("Populating Objectgegevens (Sheet8)...")
    objectcode = variables.get('object', 'UNKNOWN')
    sheet.merge_cells(start_row=21, end_row=21, start_column=4, end_column=16)
    sheet.merge_cells(start_row=24, end_row=24, start_column=4, end_column=16)
    sheet["D4"].value = "4  Objectgegevens"
    sheet["D4"].font = FONT_ARIAL_18
    sheet.row_dimensions[4].height = 25
    logging.debug("Objectgegevens populated and formatted successfully.")


def populate_risicoanalyse_sheet(
    sheet: openpyxl.worksheet.worksheet.Worksheet, variables: dict
) -> None:
    """
    Populate and format the Risicoanalyse (Sheet9).

    Parameters:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet object.
        variables (dict): Dictionary of variables.
    """
    sheet['C4'] = '5  Risicoanalyse'
    sheet['C4'].font = FONT_ARIAL_18

    # standaard tekst 5.1
    sheet["C6"] = "5.1"
    sheet["C6"].font = FONT_ARIAL_12
    sheet["D8"].font = FONT_ARIAL_10
    sheet.row_dimensions[8].height = 70

    # standaard tekst 5.2
    sheet["C10"].font = FONT_ARIAL_12
    sheet["D10"].font = FONT_ARIAL_12
    sheet["D12"].font = FONT_ARIAL_10
    sheet["D12"].alignment = ALIGNMENT_LEFT
    sheet.row_dimensions[12].height = 85

    # standaard tekst 5.3
    sheet["C14"].font = FONT_ARIAL_12
    sheet["D14"].font = FONT_ARIAL_12
    sheet["D16"].value = (
        "De inspectiemethode is opgenomen in de ORA als toevoeging op de bureaustudie,\n"
        "onderdeel 4a. Specifieke inspectie instructies, bereikbaarheidsmiddelen en \n"
        "verkeersmaatregelen zijn daarin per bouwdeel gespecificeerd. Naast specifieke\n"
        "bereikbaarheidsmiddelen heeft de IU standaard een kleine inspectieboot en ladder\n"
        "beschikbaar tijdens de inspectie. "
    )
    sheet["D16"].font = FONT_ARIAL_10
    sheet["D16"].alignment = ALIGNMENT_LEFT
    sheet.row_dimensions[16].height = 75
    logging.debug("Risicoanalyse populated and formatted successfully.")


def populate_bevindingen_sheet(
    sheet: openpyxl.worksheet.worksheet.Worksheet, variables: dict
) -> None:
    """
    Populate and format the Bevindingen (Sheet10).

    Parameters:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet object.
        variables (dict): Dictionary of variables.
    """
    logging.debug("Populating Bevindingen (Sheet10)...")
    if sheet["D5"].value == "6Bevindingen\n  \n":
        sheet["D5"] = "6  Bevindingen"
        sheet["D5"].font = FONT_ARIAL_18
        sheet["D7"].font = FONT_ARIAL_12
    else:
        sheet["E5"] = "6  Bevindingen"
        sheet["E5"].font = FONT_ARIAL_18
        sheet["E7"].font = FONT_ARIAL_12

        row_count = sheet.max_row
        for row in range(12, row_count + 1):
            text = sheet[f"I{row}"].value
            text_length = len(text) if text else 0
            sheet.row_dimensions[row].height = None
            if text_length <= 200:
                sheet.row_dimensions[row].height = 105
            else:
                factor = text_length / 200
                sheet.row_dimensions[row].height = math.ceil(factor * 105)
    logging.debug("Bevindingen populated and formatted successfully.")


def populate_bevindingenv2_sheet(
    sheet: openpyxl.worksheet.worksheet.Worksheet, variables: dict
) -> None:
    """
    Populate and format the Paragraaf (Sheet11).

    Parameters:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet object.
        variables (dict): Dictionary of variables.
    """
    logging.debug("Populating Bevindingenv2 (Sheet11)...")
    niet_schade_gerelateerd = variables.get("niet_schade_gerelateerd")
    constructieve_beoordeling = variables.get("constructieve_beoordeling")

    geen_niet_schade_gerelateerd_text = "Niet schade gerelateerde / gebruiksspecifieke risicos:\nBij dit beheerobject zijn (geen)  niet- schadegerelateerde/ gebruiksspecifieke risico’s aanwezig.\n\nVoor de volledige uitwerking en onderbouwing (inclusief eventuele schadeomschrijvingen) wordt verwezen naar Bijlage 8: Risico- en schadeomschrijving."
    wel_niet_schade_gerelateerd_text = f"Niet schade gerelateerde / gebruiksspecifieke risicos:\nBij dit beheerobject zijn de volgende niet schade gerelateerde / gebruiksspecifieke risico’s geconstateerd: {niet_schade_gerelateerd}"
    geen_constructieve_beoordeling_text = "Analyse constructieve beoordeling:\nOp basis van het objecttype is geen nadere constructieve beoordeling volgens het constructieve risico-indexerings- en afwegingsmodel (CRIAM) uitgevoerd.\n\nVoor de volledige uitwerking en onderbouwing (inclusief eventuele schadeomschrijvingen) wordt verwezen naar Bijlage 8: Risico- en schadeomschrijving"
    wel_constructieve_beoordeling_text = "Op basis van het objecttype is een nadere constructieve beoordeling volgens het constructieve risico-indexerings- en afwegingsmodel (CRIAM) uitgevoerd."

    sheet["C4"].font = FONT_ARIAL_12
    sheet["D4"].font = FONT_ARIAL_12
    sheet["D6"] = (
        wel_niet_schade_gerelateerd_text
        if niet_schade_gerelateerd
        else geen_niet_schade_gerelateerd_text
    )
    sheet["D6"].font = FONT_ARIAL_10
    sheet["C8"].font = FONT_ARIAL_12
    sheet["D8"].font = FONT_ARIAL_12
    sheet["D10"] = (
        wel_constructieve_beoordeling_text
        if constructieve_beoordeling
        else geen_constructieve_beoordeling_text
    )
    sheet.row_dimensions[6].height = 80
    sheet.row_dimensions[10].height = 80
    logging.debug("Bevindingenv2 populated and formatted successfully.")


def populate_colofon_sheet(
    sheet: openpyxl.worksheet.worksheet.Worksheet, variables: dict
) -> None:
    """
    Populate and format the Colofon (Sheet12).

    Parameters:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet object.
        variables (dict): Dictionary of variables.
    """
    logging.debug("Populating Colofon (Sheet12)...")
    opdrachtgever = variables.get("opdrachtgever", "UNKNOWN")
    contactpersoon_rws = variables.get("contactpersoon_rws", "UNKNOWN")
    zaaknummer = variables.get("zaaknummer", "UNKNOWN")
    opdrachtnemer = variables.get("opdrachtnemer", "UNKNOWN")
    contactpersoon = variables.get("projectleider", "UNKNOWN")
    projectnummer = variables.get("projectnummer", "UNKNOWN")
    code_object = variables.get("object_code", "UNKNOWN")

    sheet['C4'].font = FONT_ARIAL_18
    sheet["E7"] = f": {opdrachtgever}"
    sheet['E7'].font = FONT_ARIAL_10
    sheet["E8"] = f": {contactpersoon_rws}"
    sheet['E8'].font = FONT_ARIAL_10
    sheet["D9"] = "Zaaknummer"
    sheet["D9"].font = FONT_ARIAL_10
    sheet["E9"] = f": {zaaknummer}"
    sheet['E9'].font = FONT_ARIAL_10
    sheet["E10"] = f": {opdrachtnemer}"
    sheet['E10'].font = FONT_ARIAL_10
    sheet["E11"] = f": {contactpersoon}"
    sheet['E11'].font = FONT_ARIAL_10
    sheet["E12"] = f": {projectnummer}"
    sheet['E12'].font = FONT_ARIAL_10
    sheet["E13"] = f": PI rapport {code_object}.pdf"
    sheet['E13'].font = FONT_ARIAL_10
    logging.debug("Colofon populated and formatted successfully.")


def populate_bijlage1_sheet(
    sheet: openpyxl.worksheet.worksheet.Worksheet, variables: dict
) -> None:
    """
    Populate and format the Bijlage 1 (Sheet13).

    Parameters:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet object.
        variables (dict): Dictionary of variables.
    """
    logging.debug("Populating Bijlage 1 (Sheet13)...")
    sheet['B4'].font = FONT_ARIAL_16
    logging.debug("Bijlage 1 populated and formatted successfully.")


def populate_bijlage2_sheet(
    sheet: openpyxl.worksheet.worksheet.Worksheet, variables: dict
) -> None:
    """
    Populate and format the Bijlage 2 (Sheet14).

    Parameters:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet object.
        variables (dict): Dictionary of variables.
    """
    logging.debug("Populating Bijlage 2 (Sheet14)...")
    sheet["B4"].font = FONT_ARIAL_16
    logging.debug("Bijlage 2 populated and formatted successfully.")


def populate_bijlage3_sheet(
    sheet: openpyxl.worksheet.worksheet.Worksheet, variables: dict
) -> None:
    """
    Populate and format the Bijlage 3 (Sheet15).

    Parameters:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet object.
        variables (dict): Dictionary of variables.
    """
    logging.debug("Populating Bijlage 3 (Sheet15)...")
    sheet['B4'].font = FONT_ARIAL_16
    logging.debug("Bijlage 3 populated and formatted successfully.")


def populate_bijlage4_sheet(
    sheet: openpyxl.worksheet.worksheet.Worksheet, variables: dict
) -> None:
    """
    Populate and format the Bijlage 4 (Sheet16).

    Parameters:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet object.
        variables (dict): Dictionary of variables.
    """
    logging.debug("Populating Bijlage 4 (Sheet16)...")
    sheet['B4'].font = FONT_ARIAL_16
    row_count = sheet.max_row
    for row in range(12, row_count):
        for cell in sheet[row]:
            cell.font = FONT_ARIAL_8
    highlighted_rows = list(range(13, row_count + 5))[0::10]
    for row in highlighted_rows:
        sheet.row_dimensions[row].height = 22
    logging.debug("Bijlage 4 populated and formatted successfully.")


def _populate_bijlage5_sheet(sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """
    Populate and format the Bijlage 5 (Sheet17).

    Parameters:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet object.
    """
    logging.debug("Populating Bijlage 5 (Sheet17)...")
    sheet['B4'].font = FONT_ARIAL_16
    if sheet["C6"].value:
        sheet["C6"] = "Omgevingsfoto schade"
        sheet["C6"].font = FONT_ARIAL_10_BOLD
        sheet["D6"] = "Schadefoto"
        sheet["D6"].font = FONT_ARIAL_10_BOLD

        # sheet.unmerge_cells(start_row=7, start_column=4, end_row=8, end_column=7)
        # sheet.merge_cells("D7:G7")
        beschrijving = sheet["C8"].value
        sheet["C10"].value = beschrijving
        sheet["C8"].value = ""
        sheet["C10"].font = FONT_ARIAL_10
        sheet["C10"].alignment = ALIGNMENT_LEFT
        sheet.row_dimensions[10].height = 300
        sheet.row_dimensions[10].hidden = False
        # sheet.row_dimensions[7].height = 300
        sheet.row_dimensions[9].height = 2
    logging.debug("Bijlage 5 populated and formatted successfully.")


def populate_bijlage5_plus_return_next_idx(wb: openpyxl.Workbook) -> int:
    """
    Populate and format additional Bijlage 5 sheets.

    Parameters:
        wb (openpyxl.Workbook): The workbook object.
        sheets_count (int): Total number of sheets.
    """
    logging.debug("Populating additional Bijlage 5 sheets...")
    _populate_bijlage5_sheet(wb["Sheet17"])
    # Get all sheet names from the workbook
    sheet_names = wb.sheetnames
    sheets_count = len(sheet_names)

    # Find the starting index for "Sheet 18"
    start_index = sheet_names.index("Sheet18")
    last_processed_index = None

    # Loop through the sheets starting from "Sheet 18"
    for i in range(start_index, sheets_count):
        sheet = wb[sheet_names[i]]

        # Check if cell B4 contains "Omgevingsfoto schade"
        if sheet["C4"].value == "Omgevingsfoto schade":
            # Perform the required operations
            if sheet.row_dimensions[8].height >= 5:
                text_2 = sheet["C6"].value
                sheet["C8"].value = text_2
                sheet["C6"].value = ""
                sheet["C8"].font = FONT_ARIAL_10
                sheet["C8"].alignment = ALIGNMENT_LEFT
                sheet.row_dimensions[8].height = 300
                sheet.row_dimensions[8].hidden = False
                sheet.row_dimensions[7].height = 2
            else:
                sheet.row_dimensions[6].height = 300
                sheet["C6"].font = FONT_ARIAL_10
            # Update the last processed sheet index
            last_processed_index = i
        else:
            break

    # Determine the sheet that comes after the last processed sheet
    if last_processed_index is not None and last_processed_index + 1 < len(sheet_names):
        next_sheet_idx = int(last_processed_index + 1)
        logging.debug("Additional Bijlage 5 sheets populated successfully.")
        return next_sheet_idx
    else:
        # If no sheets were processed or the last processed sheet is the last one
        logging.debug("No additional Bijlage 5 sheets to populate.")
        return start_index


def populate_bijlage6_sheet(
    sheet: openpyxl.worksheet.worksheet.Worksheet, variables: dict
) -> None:
    """
    Populate and format the Bijlage 6 sheet.

    Parameters:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet object.
        variables (dict): Dictionary of variables.
    """
    logging.debug("Populating Bijlage 6 (Sheet)...")
    # inspectietekeningen = variables.get("inspectietekeningen")
    sheet["B4"].font = FONT_ARIAL_16
    # if inspectietekeningen.lower() == "ja":
    sheet["C6"].value = ""
    logging.debug("Bijlage 6 populated and formatted successfully.")


def populate_bijlage7_sheet(
    sheet: openpyxl.worksheet.worksheet.Worksheet, variables: dict
) -> None:
    """
    Populate and format the Bijlage 7 sheet.

    Parameters:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet object.
        variables (dict): Dictionary of variables.
    """
    logging.debug("Populating Bijlage 7 (Sheet)...")
    sheet['C4'].font = FONT_ARIAL_16
    logging.debug("Bijlage 7 populated and formatted successfully.")


def populate_bijlage8_sheet(
    sheet: openpyxl.worksheet.worksheet.Worksheet, variables: dict
) -> None:
    """
    Populate and format the Bijlage 8 sheet.

    Parameters:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet object.
        variables (dict): Dictionary of variables.
    """
    logging.debug("Populating Bijlage 8 (Sheet)...")
    sheet['C5'].font = FONT_ARIAL_16
    logging.debug("Bijlage 8 populated and formatted successfully.")


def populate_bijlage8_2_sheet(
    sheet: openpyxl.worksheet.worksheet.Worksheet, variables: dict
) -> None:
    """
    Populates the specified Excel sheet with data for 'Bijlage 8.2'.
    Args:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The Excel worksheet to populate.
        variables (dict): A dictionary containing the variables to be used in the sheet.
            Expected keys:
                - 'opdrachtnemer' (str): The name of the contractor. Defaults to 'UNKNOWN' if not provided.
    Modifies:
        The function modifies the following cells and properties in the sheet:
            - Sets the height of row 7 to 45.
            - Sets the value of cell F23 to the 'opdrachtnemer' value.
            - Sets the height of row 18 to 30.
    """
    logging.debug("Populating Bijlage 8.2 (Sheet)...")
    opdrachtnemer = variables.get("opdrachtnemer", "UNKNOWN")
    if str(sheet["F20"].value).endswith("(IN_UITVOERING)"):
        sheet["F20"] = str(sheet["F20"].value).replace(
            "(IN_UITVOERING)", "(VASTGESTELD)"
        )
    sheet.row_dimensions[7].height = 45
    sheet["F23"] = opdrachtnemer
    sheet.row_dimensions[18].height = 30
    logging.debug("Bijlage 8.2 populated and formatted successfully.")


def populate_bijlage8_3_sheet(
    sheet: openpyxl.worksheet.worksheet.Worksheet, variables: dict
) -> None:
    """
    Populates the specified Excel sheet with data for 'Bijlage 8.3'.
    Args:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The Excel worksheet to populate.
        variables (dict): A dictionary containing the variables to be used in the sheet.
    """
    logging.debug("Populating Bijlage 8.3 (Sheet)...")
    max_row_count = sheet.max_row
    start_sheet = 4
    combined_range = [
        f"{cell[0].value or ''}{cell[1].value or ''}"
        for cell in sheet.iter_rows(
            min_row=start_sheet, max_row=max_row_count, min_col=3, max_col=4
        )
    ]

    empty_rows = [
        idx + start_sheet for idx, value in enumerate(combined_range) if not value
    ]
    filled_rows = [
        idx + start_sheet for idx, value in enumerate(combined_range) if value
    ]

    end_range = 6

    # Toestand Karakteristiek
    tables = {}
    while True:
        # Find the next filled row after the current end_range
        try:
            start_range = next(row for row in filled_rows if row > end_range)
        except StopIteration:
            break

        table_name = sheet[f"C{start_range}"].value

        # Find the next empty row after start_range
        next_empty = next((row for row in empty_rows if row > start_range), None)

        # Find the next filled row after next_empty to set end_range
        if next_empty:
            end_range = next(
                (row for row in filled_rows if row > next_empty), next_empty
            )
            end_range -= 1
        else:
            end_range = max_row_count + 1

        # Collect all rows from start_range to end_range - 1
        if not table_name:
            continue
        table_name = table_name + str(start_range)
        tables[table_name] = list(range(start_range, end_range))

    for name, rows in tables.items():
        if name.startswith("Toestand karakteristiek"):
            for row in rows[:-1]:
                sheet.row_dimensions[row].height = 15
        elif name.startswith("IH-onderdeelnaam"):
            for row in rows[1:]:
                sheet.row_dimensions[row].height = 45
                # TODO: Check waarom deze niet goed gaat
        elif name.startswith("Aspecteis"):
            for row in rows[1:]:
                sheet.row_dimensions[row].height = 130
        elif name.startswith("Afgemelde"):
            if len(rows) > 2:
                for row in rows[2:]:
                    sheet.row_dimensions[row].height = 45
        elif (
            name and not name.startswith("Referentiegegevens") and "-" in name
        ):  # Dit zijn de specifieke elementen, zoals hoofddraagconstructie
            for row in [rows[0] + idx for idx in [7, 9, 18, 21]]:
                sheet.row_dimensions[row].height = 80
            # Dit zijn de specifieke elementen
    logging.debug("Bijlage 8.3 populated and formatted successfully.")


def populate_bijlage9_sheet(
    sheet: openpyxl.worksheet.worksheet.Worksheet, variables: dict
) -> None:
    """
    Populate and format the Bijlage 9 sheet.

    Parameters:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet object.
        variables (dict): Dictionary of variables.
    """
    logging.debug("Populating Bijlage 9 (Sheet)...")
    sheet['C5'].font = FONT_ARIAL_16
    logging.debug("Bijlage 9 populated and formatted successfully.")


def populate_bijlage10_sheet(
    sheet: openpyxl.worksheet.worksheet.Worksheet, variables: dict
) -> None:
    """
    Populate and format the Bijlage 10 sheet.

    Parameters:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet object.
        variables (dict): Dictionary of variables.
    """
    logging.debug("Populating Bijlage 10 (Sheet)...")
    criam = variables.get('criam', 'UNKNOWN')

    sheet['C5'].font = FONT_ARIAL_16
    sheet["D7"].font = FONT_ARIAL_10
    logging.debug("Bijlage 10 populated and formatted successfully.")


def update_config_variables(
    sheet: openpyxl.worksheet.worksheet.Worksheet, variables: dict
) -> dict:
    """
    Update the config variables.

    Parameters:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet object.
        variables (dict): Dictionary of variables.

    Returns:
        dict: Updated dictionary of variables.
    """
    logging.info("Updating variables with config variables...")
    # Find the input list workbook
    config_variables = variables.copy()

    config_variables["object_omschrijving"] = sheet["H8"].value
    config_variables["object_naam"] = sheet["H9"].value
    config_variables["object_beheer"] = sheet["H10"].value
    current_date = dt.datetime.now().strftime("%d-%m-%Y")
    config_variables["datum"] = (
        current_date
        if pd.isna(config_variables["datum"])
        else config_variables["datum"]
    )
    logging.info("Config variables updated successfully.")
    return config_variables



def process_pi_report_for_object(
    object_path: str, report_path: str, config: dict
) -> None:
    """
    Process the PI report for a specific object.

    Parameters:
        object_path (str): Path to the object.
        report_path (str): Path to the (original) report.
        config (dict): Configuration dictionary.
    """
    logging.info(f"Processing PI report for [{report_path}]")

    # Load workbooks
    wb_report = utilsxls.load_workbook(report_path)
    sheet_names = wb_report.sheetnames

    # Delete mpo images
    mpo_images = utilsxls.find_mpo_references(wb_report)
    logging.info(f"Found `.mpo` images: {mpo_images}")

    # Step 2: Replace `.mpo` images with `.png`
    if mpo_images:
        utilsxls.delete_images(wb_report, mpo_images)
        logging.info(f"Deleted `.mpo` images: {mpo_images} for object {object_path}")

    # update variables
    config_variables = update_config_variables(wb_report["Sheet2"], config)

    # Populate excel with variables
    populate_title_page(wb_report["Sheet2"], config_variables)
    populate_inhoud_sheet(wb_report["Sheet3"], config_variables)
    populate_aanbeveling_sheet(wb_report["Sheet4"], config_variables)
    populate_ihp_sheet(wb_report["Sheet5"], config_variables)
    populate_miok_sheet(wb_report["Sheet6"], config_variables)
    populate_inleiding_sheet(wb_report["Sheet7"], config_variables)
    populate_objectgegevens_sheet(wb_report["Sheet8"], config_variables)
    populate_risicoanalyse_sheet(wb_report["Sheet9"], config_variables)
    populate_bevindingen_sheet(wb_report["Sheet10"], config_variables)
    populate_bevindingenv2_sheet(wb_report["Sheet11"], config_variables)
    populate_colofon_sheet(wb_report["Sheet12"], config_variables)
    populate_bijlage1_sheet(wb_report["Sheet13"], config_variables)
    populate_bijlage2_sheet(wb_report["Sheet14"], config_variables)
    populate_bijlage3_sheet(wb_report["Sheet15"], config_variables)
    populate_bijlage4_sheet(wb_report["Sheet16"], config_variables)
    next_sheet_idx = populate_bijlage5_plus_return_next_idx(wb_report)
    populate_bijlage6_sheet(
        wb_report[sheet_names[next_sheet_idx]], config_variables
    )  # TODO: Hier moet een check komen in voortgangslijst of er inspectietekeningen zijn.

    populate_bijlage7_sheet(
        wb_report[sheet_names[next_sheet_idx + 1]], config_variables
    )
    populate_bijlage8_sheet(
        wb_report[sheet_names[next_sheet_idx + 4]], config_variables
    )
    populate_bijlage8_2_sheet(
        wb_report[sheet_names[next_sheet_idx + 5]], config_variables
    )
    populate_bijlage8_3_sheet(
        wb_report[sheet_names[next_sheet_idx + 6]], config_variables
    )
    populate_bijlage9_sheet(
        wb_report[sheet_names[next_sheet_idx + 7]], config_variables
    )
    populate_bijlage10_sheet(
        wb_report[sheet_names[next_sheet_idx + 8]], config_variables
    )
    set_footer(
        wb_report, wb_report.sheetnames, config_variables, len(wb_report.sheetnames)
    )
    logging.info("Finished populating the PI report.")

    # Save the workbook
    utilsxls.save_and_finalize_workbook(wb_report, config_variables, save_dir=object_path)

    logging.info(f"Done for {config_variables['object_code']}")


def print_excel_to_pdf(path_of_pi_report: str) -> None:
    """
    Print the Excel report to PDF.

    Parameters:
        path_of_pi_report (str): Path to the PI report.
    """
    run_macro_on_workbook(path_of_pi_report, "InspectieRapportage", "ExportToPDF")
    logging.info("PI report processing completed successfully.")


def main() -> None:
    """
    Main function to orchestrate the processing of the PI report.
    """
    logger = utils.setup_logger("generate_pi_report.log", log_level="INFO")
    config = utils.load_config('./config.json')
    logger.info(f"Starting PI report processing for werkpakket [{config['werkpakket']}]")
    failed_objects = []

    # Get the voortgangs data, based on the excel file (as set in config.json)
    excelfile = config.get("voortgangs_sheet", "")
    voortgangs_data = get_voortgang(excelfile)
    
    for object_path, object_code in utils.get_object_paths_codes():
        try:
            logger.info(f"Processing object [{object_code}]")
            logger.info(f"Updating the configuration variables with voortgang...")
            voortgang = get_voortgang_params(voortgangs_data, object_code)
            config = utils.update_config_with_voortgang(config, voortgang)
            pi_report_path = find_inspectierapport(object_path)
            if not pi_report_path:
                logger.error(f"Could not find inspectierapport for [{object_code}]")
                raise FileNotFoundError(f"Could not find inspectierapport for [{object_code}]")
            process_pi_report_for_object(object_path, pi_report_path, config)
            
            # print_excel_to_pdf(os.path.join(save_loc, f"PI rapport {object_code}.xlsx"))
        except Exception as e:
            logger.error(f"Error processing [{object_code}]: {e}")
            failed_objects.append(object_code)
    if failed_objects:
        logger.error(f"Failed to process the following objects: {failed_objects}")
    else:
        logger.info("All objects processed successfully.")


if __name__ == "__main__":
    main()
