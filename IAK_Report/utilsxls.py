# IAK Reporting Tool - Excel utilities
# Copyright (C) 2024-2025 Arcadis Nederland B.V.
#
# SPDX-License-Identifier: GPL-3.0-or-later
# See LICENSE file for full license text.

"""
This module contains utility functions for handling Excel files with openpyxl.

It includes functionalities to:

- Find the most recently modified ORA file in a directory.
- Find and delete references to images with `.mpo` extensions in the workbook.
"""

# Built-in imports
import os
import logging

# External imports
import pandas as pd
import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
import win32com.client


def load_workbook(path: str) -> openpyxl.Workbook:
    """
    Load an Excel workbook, returning the workbook object.

    Parameters:
        path (str): Full path to the Excel workbook.

    Returns:
        openpyxl.Workbook: Loaded workbook object.
    """
    try:
        logging.debug(f"Loading Excel workbook from [{path}]...")
        wb = openpyxl.load_workbook(path)  #, rich_text=True)
        logging.debug("Workbook loaded successfully.")
        return wb
    except FileNotFoundError:
        logging.error(f"Error: The file at [{path}] was not found.")
        raise
    except Exception as e:
        logging.error(f"An unexpected error occurred while loading the workbook: {e}")
        raise



def styling_cell_with_colons(plain_text: str) -> openpyxl.cell.rich_text.CellRichText:
    """
    Convert plain text with colons into rich text blocks for Excel cells.

    Parameters:
        plain_text (str): Text to be converted, where colons indicate bold text.

    Returns:
        openpyxl.cell.rich_text.CellRichText: Rich text object ready for Excel cell.
    """
    

    rich_text = CellRichText()
    for line in plain_text.splitlines():
        if ':' in line:
            before_colon, after_colon = line.split(':', 1)
            rich_text.append(TextBlock(text=before_colon + ':', font=InlineFont(b=True)))
            rich_text.append(TextBlock(text=after_colon + '\n', font=InlineFont(b=False)))
        else:
            rich_text.append(TextBlock(text=line, font=InlineFont(b=False)))
    
    return rich_text


def find_ora_sheet_name(workbook) -> str | None:
    """
    Find the sheet in the workbook that corresponds to the ORA report.

    Args:
        workbook: The openpyxl workbook object.
        or
        str: The path to the Excel workbook.
          (could be an xlsb, which openpyxl can't read, but pandas does)

    Returns:
        The sheet name if found, otherwise None.
    """
    
    # If a string is given, open this path, and get the workbook object
    if isinstance(workbook, str):
        list_sheets = pd.ExcelFile(workbook).sheet_names
    else:
        list_sheets = workbook.sheetnames

    for sheet in list_sheets:
        if sheet.startswith("ORA"):
            return sheet
    return None


def find_mpo_references(workbook):
    """
    Find references to images with `.mpo` extensions in the workbook.
    MPO references lead to problems in saving the workbook.

    Args:
        workbook: The openpyxl workbook object.

    Returns:
        A list of tuples containing (sheet_name, image_reference).
    """
    mpo_references = []
    for sheet in workbook.worksheets:
        if hasattr(sheet, "_images"):  # Check if the sheet contains images
            for img in sheet._images:
                # Check if the image has a path attribute and contains '.mpo'
                if hasattr(img, "path") and ".mpo" in img.path:
                    mpo_references.append((sheet.title, img))
    return mpo_references


def delete_images(workbook, image_references):
    """
    Delete images from the workbook based on the provided references.
    I was unable to find a way to convert them. I also cannot find the mpo files in the underlying zip file.

    Args:
        workbook: The openpyxl workbook object.
        image_references: A list of tuples containing (sheet_name, image_reference).

    Returns:
        None
    """
    for sheet_name, img in image_references:
        # Get the sheet by name
        sheet = workbook[sheet_name]  
        if hasattr(sheet, "_images"):  # Ensure the sheet has images
            # Check if the image is in the sheet's images
            if img in sheet._images:  
                sheet._images.remove(img)  # Remove the image
                logging.info(f"Deleted image: {img.path} from sheet: {sheet_name}")


def save_and_finalize_workbook(wb: openpyxl.Workbook, variables: dict, save_dir: str) -> str:
    """
    Save and finalize the workbook.

    Parameters:
        wb (openpyxl.Workbook): Workbook to save.
        variables (dict): Dictionary of variables.
        target_dir (str): Directory to save the workbook.
    """
    object_code = variables.get("object_code", "UNKNOWN")
    filename_excel = f"PI rapport {object_code}.xlsx"
    filepath_excel = os.path.join(save_dir, filename_excel)

    # Create save directory if it doesn't exist
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)
        logging.info(f"Created directory: {save_dir}")

    # Remove 'Document map' sheet if it exists
    if "Document map" in wb.sheetnames:
        del wb["Document map"]

    # Set active sheet to 'Sheet2'
    wb.active = wb["Sheet2"]

    # Ensure all sheets are marked as selected
    for sheet in wb:
        sheet.views.sheetView[0].tabSelected = True

    # Save the workbook
    logging.debug(f"Saving workbook to {filepath_excel}...")
    wb.save(filepath_excel)
    logging.info(f"Workbook saved: {filename_excel}")

    return filepath_excel


def styling_bijlage3_export(worksheet, excel: win32com.client.Dispatch) -> None:
    """
    Apply styling to the export of Bijlage 3.

    Args:
        worksheet (Worksheet): The worksheet to style.
    """

    # Hide some columns
    worksheet.Range("K5").Value = "1.0 - Definitief"
    worksheet.Columns("AJ:BI").Hidden = True
    worksheet.Columns("CB:CG").Hidden = True

    # Set margins
    worksheet.PageSetup.TopMargin = excel.Application.CentimetersToPoints(1.91)
    worksheet.PageSetup.BottomMargin = excel.Application.CentimetersToPoints(1.91)
    worksheet.PageSetup.LeftMargin = excel.Application.CentimetersToPoints(0.64)
    worksheet.PageSetup.RightMargin = excel.Application.CentimetersToPoints(0.64)
    worksheet.PageSetup.HeaderMargin = excel.Application.CentimetersToPoints(0.76)
    worksheet.PageSetup.FooterMargin = excel.Application.CentimetersToPoints(0.76)

    # Set title rows and print area
    worksheet.PageSetup.FitToPagesWide = 1
    worksheet.PageSetup.FitToPagesTall = False
    worksheet.PageSetup.PrintTitleRows = "$8:$11"
    worksheet.PageSetup.PrintArea = "A:CZ"

    return None  # Modified worksheet in place


def export_to_pdf(excel_path: str, pdf_path: str, sheet_name: str = None) -> None:
    """
    Export an Excel file to PDF using Excel's built-in functionality via COM automation (Windows only).
    This emulates the steps "File" > "Export" > "Create PDF/XPS Document".

    Parameters:
        excel_path (str): Path to the Excel file.
        pdf_path (str): Path where the PDF should be saved.
        sheet_name (str, optional): Name of the sheet to export. If None, all sheets are exported.

    Raises:
        RuntimeError: If export fails.
    """
    # Create save directory if it doesn't exist
    save_dir = os.path.dirname(pdf_path)
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)
        logging.info(f"Created directory: {save_dir}")
    try:
        import win32com.client
        
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False  # Suppress pop-up alerts
        wb = excel.Workbooks.Open(excel_path)
        if sheet_name:
            # Export the specified sheet
            ws = wb.Worksheets(sheet_name)
            # Set the page styling before printing
            # Not generic, but for Bijlage 3 it will do
            styling_bijlage3_export(ws, excel)

            ws.ExportAsFixedFormat(0, pdf_path)  # 0 = PDF
        else:
            # Export all sheets
            wb.ExportAsFixedFormat(0, pdf_path)  # 0 = PDF
        wb.Close(False)
        excel.Quit()
    except Exception as e:
        raise RuntimeError(f"Failed to export Excel to PDF: {e}")
