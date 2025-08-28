import datetime as dt

import openpyxl
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import CellRichText, TextBlock

# Create a new workbook and select the active worksheet
wb = openpyxl.Workbook()
ws = wb.active

# Define the lines for rich text
text = """
Word1: This is the description for word one.
Word2: This is the description for word two.
Word3: This is the description for word three.
"""

# Build rich text blocks
rich_text_blocks = []
for line in text.strip().split("\n"):
    word, desc = line.split(": ", 1)
    # Bold for word and colon
    rich_text_blocks.append(TextBlock(text=word + ": ", font=InlineFont(b=True)))
    # Regular for description and newline
    rich_text_blocks.append(desc + "\n")
    
# Assign rich text to cell A1
ws["A1"].value = CellRichText(rich_text_blocks)
ws["A1"].alignment = openpyxl.styles.Alignment(wrap_text=True)

# Save the workbook with a name based on the current time 
# (to avoid permission errors on open files)
timestamp = dt.datetime.now().strftime("%Y%m%dT%H%M%S")
wb.save(f"test_rich_text_{timestamp}.xlsx")
